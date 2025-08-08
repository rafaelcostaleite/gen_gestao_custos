# src/utils/excel_exporter.py

import pandas as pd
import logging
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelExporter:
    @staticmethod
    def export_with_formatting(df: pd.DataFrame, output_path: Path, 
                             sheet_name: str, numeric_columns: list):
        """Exporta DataFrame com formatação profissional"""
        logging.info(f"Exportando para: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            ExcelExporter._adjust_column_widths(worksheet)
            ExcelExporter._format_header(worksheet)
            ExcelExporter._format_numeric_columns(worksheet, numeric_columns)
        
        logging.info("Formatação aplicada com sucesso")
    
    @staticmethod
    def _adjust_column_widths(worksheet):
        """Ajusta largura das colunas automaticamente"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _format_header(worksheet):
        """Aplica formatação ao cabeçalho"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    @staticmethod
    def _format_numeric_columns(worksheet, numeric_columns):
        """Formata colunas numéricas"""
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column_letter in numeric_columns:
                    try:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    except:
                        pass
                
                cell.alignment = Alignment(horizontal="left", vertical="center")